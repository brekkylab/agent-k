import math
import os
import re
import shutil
import subprocess
import sys
import tempfile
import unicodedata
from datetime import date, datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


ISO_DATE_PATTERN = re.compile(r"^\d{4}-\d{2}-\d{2}$")

ERROR_VALUES = {"#NAME?", "#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NULL!", "#NUM!"}


def _visual_text_width(s):
    """Display width of a string: CJK / full-width chars count as 2, others 1."""
    return sum(2 if unicodedata.east_asian_width(ch) in ("W", "F") else 1 for ch in s)


def _recalc_workbook(path):
    """Run the file through LibreOffice headless (evaluating every formula)
    and return a data_only Workbook with cached results."""
    soffice = shutil.which("libreoffice") or shutil.which("soffice")
    if not soffice:
        raise RuntimeError("LibreOffice not found (need libreoffice/soffice on PATH)")
    with tempfile.TemporaryDirectory() as tmpdir:
        subprocess.run(
            [soffice, "--headless", "--calc", "--convert-to", "xlsx",
             "--outdir", tmpdir, str(path)],
            check=True, capture_output=True, timeout=120,
        )
        recalc = os.path.join(tmpdir, os.path.basename(path))
        return load_workbook(recalc, data_only=True)


def verify_formulas(path):
    """Recalculate all formulas via LibreOffice and return any error cells.

    openpyxl stores formula strings without evaluating them — you can't tell
    if `=RANK.EQ(...)` will show #NAME? until Excel/LibreOffice computes it.
    Run it as the last step after `save()`.

    Returns: list of (sheet_title, coordinate, error_string).
             Empty list means no formula errors.
    Raises RuntimeError if LibreOffice is not on PATH.
    """
    wb = _recalc_workbook(path)
    errors = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value in ERROR_VALUES:
                    errors.append((ws.title, cell.coordinate, cell.value))
    return errors


def read_computed(path, cells=None):
    """Recalculate via LibreOffice and return computed cell *values* — for
    sanity-checking results against expectations (e.g. is L4 really
    supply + fee + VAT? does the column total match?).

    `cells`: optional dict {sheet_title: [coord, ...]} to fetch only specific
    cells. If None, returns every non-empty cell of every sheet.

    Returns: dict {sheet_title: {coord: value}} with computed values.
    Raises RuntimeError if LibreOffice is not on PATH.
    """
    wb = _recalc_workbook(path)
    out = {}
    if cells is not None:
        for sheet, coords in cells.items():
            ws = wb[sheet]
            out[sheet] = {c: ws[c].value for c in coords}
        return out
    for ws in wb.worksheets:
        out[ws.title] = {
            cell.coordinate: cell.value
            for row in ws.iter_rows()
            for cell in row
            if cell.value is not None
        }
    return out


def _display_text(value, number_format):
    """Approximate the string Excel actually shows for a value under its
    number_format — so column widths reflect the *displayed* text, not the
    raw formula or raw number."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (date, datetime)):
        fmt = number_format or ""
        return fmt if fmt and fmt != "General" else str(value)
    if isinstance(value, (int, float)):
        fmt = number_format or ""
        m = re.search(r"\.(0+)", fmt)          # decimals after the dot
        decimals = len(m.group(1)) if m else 0
        is_pct = "%" in fmt
        v = value * 100 if is_pct else value
        comma = ("#,#" in fmt) or ("," in fmt)
        s = f"{v:,.{decimals}f}" if comma else f"{v:.{decimals}f}"
        for sym in ("₩", "$", "€", "¥", "£"):
            if sym in fmt:
                s = sym + s
                break
        if is_pct:
            s += "%"
        return s
    return str(value)


def autofit_columns(path, header_row=3):
    """Resize column widths from *actual computed values* (via LibreOffice).

    openpyxl can't see formula results, so a currency column full of `=I*0.2`
    gets sized to a guess and shows `######`. This recalculates via
    LibreOffice, measures each column by its real displayed value
    (number-format aware for numbers, east-asian-width aware for CJK text),
    and rewrites the widths. Run after `save()` when columns hold formulas or
    formula-returned CJK text. Re-saves `path` in place.
    """
    computed = _recalc_workbook(path)
    wb = load_workbook(path)
    for ws in wb.worksheets:
        cws = computed[ws.title] if ws.title in computed.sheetnames else None
        for col_cells in ws.columns:
            col_idx = col_cells[0].column
            max_w = 0
            for cell in col_cells:
                if cell.row < header_row:
                    continue
                if cell.row == header_row or cws is None:
                    disp = "" if cell.value is None else str(cell.value)
                else:
                    cv = cws.cell(row=cell.row, column=col_idx).value
                    disp = _display_text(cv, cell.number_format)
                w = _visual_text_width(disp)
                if w > max_w:
                    max_w = w
            if max_w > 0:
                ws.column_dimensions[get_column_letter(col_idx)].width = max(
                    XLSXReportSkill.COLUMN_WIDTH_MIN,
                    min(max_w + XLSXReportSkill.COLUMN_WIDTH_PADDING,
                        XLSXReportSkill.COLUMN_WIDTH_MAX),
                )
    wb.save(path)
    return path


class Formula:

    @staticmethod
    def sum(range_ref):
        return f"=SUM({range_ref})"

    @staticmethod
    def average(range_ref):
        return f"=AVERAGE({range_ref})"

    @staticmethod
    def safe_divide(a, b):
        return f"=IFERROR({a}/{b}, 0)"

    @staticmethod
    def rank(cell, range_ref, descending=True):
        """Classic RANK (pre-2010); safe without prefix."""
        order = 0 if descending else 1
        return f"=RANK({cell},{range_ref},{order})"

    @staticmethod
    def rank_eq(cell, range_ref, descending=True):
        order = 0 if descending else 1
        return f"=_xlfn.RANK.EQ({cell},{range_ref},{order})"

    @staticmethod
    def rank_avg(cell, range_ref, descending=True):
        order = 0 if descending else 1
        return f"=_xlfn.RANK.AVG({cell},{range_ref},{order})"

    @staticmethod
    def ifs(*conditions_and_results):
        """Compiles to nested IF; last "TRUE" condition becomes the else-branch."""
        if len(conditions_and_results) == 0 or len(conditions_and_results) % 2 != 0:
            raise ValueError(
                "Formula.ifs() requires an even, non-zero number of args "
                "(condition, result, condition, result, ...)"
            )
        pairs = list(zip(conditions_and_results[::2], conditions_and_results[1::2]))
        if str(pairs[-1][0]).strip().upper() == "TRUE":
            else_branch = str(pairs[-1][1])
            pairs = pairs[:-1]
        else:
            else_branch = '""'
        expr = else_branch
        for cond, val in reversed(pairs):
            expr = f"IF({cond},{val},{expr})"
        return f"={expr}"

    @staticmethod
    def xlookup(lookup_value, lookup_array, return_array, if_not_found='""'):
        return (
            f"=_xlfn.XLOOKUP({lookup_value},{lookup_array},"
            f"{return_array},{if_not_found})"
        )

    @staticmethod
    def maxifs(max_range, *criteria_pairs):
        parts = ",".join(str(x) for x in criteria_pairs)
        return f"=_xlfn.MAXIFS({max_range},{parts})"

    @staticmethod
    def minifs(min_range, *criteria_pairs):
        parts = ",".join(str(x) for x in criteria_pairs)
        return f"=_xlfn.MINIFS({min_range},{parts})"

    @staticmethod
    def ifna(value, if_na_value):
        return f"=_xlfn.IFNA({value},{if_na_value})"

    @staticmethod
    def concat(*range_refs):
        parts = ",".join(str(x) for x in range_refs)
        return f"=_xlfn.CONCAT({parts})"

    @staticmethod
    def sumifs_month(sum_range, date_range, year, month):
        anchor = f"DATE({year},{month},1)"
        return (
            f"=SUMIFS({sum_range},{date_range},\">=\"&{anchor},"
            f"{date_range},\"<=\"&EOMONTH({anchor},0))"
        )

    @staticmethod
    def sumifs_date_range(sum_range, date_range, start_date, end_date):
        """start_date / end_date are (year, month, day) tuples."""
        sy, sm, sd = start_date
        ey, em, ed = end_date
        return (
            f"=SUMIFS({sum_range},{date_range},\">=\"&DATE({sy},{sm},{sd}),"
            f"{date_range},\"<=\"&DATE({ey},{em},{ed}))"
        )


class XLSXReportSkill:

    TITLE_FILL = PatternFill(start_color="1F1F1F", end_color="1F1F1F", fill_type="solid")
    HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ROW_FILL = PatternFill(start_color="F4F6F8", end_color="F4F6F8", fill_type="solid")
    WHITE_FONT = Font(bold=True, color="FFFFFF")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    NORMAL_FONT = Font(color="000000")
    BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ROW_HEIGHT_TITLE = 28
    ROW_HEIGHT_HEADER = 22
    ROW_HEIGHT_DATA = 20
    ROW_HEIGHT_SPACER = 10

    COLUMN_WIDTH_MIN = 10
    COLUMN_WIDTH_MAX = 40
    COLUMN_WIDTH_PADDING = 4
    COLUMN_WIDTH_FORMULA_DEFAULT = 8

    def __init__(self):
        self.wb = Workbook()
        self.wb.remove(self.wb.active)

    # ----- SHEET -----

    def add_sheet(self, name, title, headers, rows):
        ws = self.wb.create_sheet(title=name)
        self.render_layout(ws, title, headers, rows)
        self.apply_style(ws)
        self.configure_sheet(ws)

    # ----- LAYOUT -----

    def render_layout(self, ws, title, headers, rows):
        total_columns = len(headers)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
        ws.cell(row=1, column=1).value = title

        for col_index, header in enumerate(headers, start=1):
            ws.cell(row=3, column=col_index).value = header

        for row_index, row_data in enumerate(rows):
            for col_index, value in enumerate(row_data, start=1):
                if isinstance(value, str) and ISO_DATE_PATTERN.match(value):
                    try:
                        value = datetime.strptime(value, "%Y-%m-%d").date()
                    except ValueError:
                        pass

                cell = ws.cell(row=4 + row_index, column=col_index)
                cell.value = value
                if isinstance(value, (date, datetime)):
                    cell.number_format = "yyyy-mm-dd"

    # ----- STYLE -----

    def apply_style(self, ws):
        for cell in ws[1]:
            cell.fill = self.TITLE_FILL
            cell.font = self.WHITE_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws[3]:
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.BORDER
            cell.alignment = Alignment(horizontal="center")

        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = self.NORMAL_FONT
                cell.border = self.BORDER
                if cell.row % 2 == 0:
                    cell.fill = self.ROW_FILL

    # ----- CONFIG -----

    @classmethod
    def _data_end_row(cls, ws, header_row=3):
        """Last row of the contiguous data block; stops at the first all-blank row."""
        last_data = header_row
        for r in range(header_row + 1, ws.max_row + 1):
            if all(ws.cell(row=r, column=c).value is None for c in range(1, ws.max_column + 1)):
                break
            last_data = r
        return last_data

    @classmethod
    def _compute_column_visual_width(cls, column_cells, header_row=3, last_data_row=None):
        max_length = 0
        for cell in column_cells:
            if cell.row < header_row:
                continue
            if last_data_row is not None and cell.row > last_data_row:
                continue
            value = cell.value
            if value is None:
                continue
            s = str(value)
            length = cls.COLUMN_WIDTH_FORMULA_DEFAULT if s.startswith("=") else _visual_text_width(s)
            if length > max_length:
                max_length = length
        return max_length

    def configure_sheet(self, ws):
        ws.freeze_panes = None  # don't use freeze panes

        last_data_row = self._data_end_row(ws)
        ws.auto_filter.ref = None  # don't use auto filter

        for column_cells in ws.columns:
            max_length = self._compute_column_visual_width(column_cells, last_data_row=last_data_row)
            adjusted_width = max(
                self.COLUMN_WIDTH_MIN,
                min(max_length + self.COLUMN_WIDTH_PADDING, self.COLUMN_WIDTH_MAX),
            )
            letter = get_column_letter(column_cells[0].column)
            ws.column_dimensions[letter].width = adjusted_width

        # Scale each row's height by its tallest multi-line cell so wrapped
        # text (e.g. a 2-line header like "Opening\n(2024-05)") is not clipped.
        ws.row_dimensions[1].height = self.ROW_HEIGHT_TITLE * self._row_line_count(ws, 1)
        ws.row_dimensions[2].height = self.ROW_HEIGHT_SPACER
        ws.row_dimensions[3].height = self.ROW_HEIGHT_HEADER * self._row_line_count(ws, 3)
        for r in range(4, ws.max_row + 1):
            ws.row_dimensions[r].height = self.ROW_HEIGHT_DATA * self._row_line_count(ws, r)

    @staticmethod
    def _row_line_count(ws, r):
        """Max newline-separated line count among a row's cells (>= 1), so a
        wrapped multi-line header/label gets a row tall enough to show it."""
        n = 1
        for c in ws[r]:
            v = c.value
            if isinstance(v, str) and "\n" in v:
                n = max(n, v.count("\n") + 1)
        return n

    # ----- DEFINED NAMES (named ranges) -----

    _ASCII_SHEET_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")

    @classmethod
    def quote_sheet_name(cls, sheet_name):
        if cls._ASCII_SHEET_RE.match(sheet_name):
            return sheet_name
        return f"'{sheet_name.replace(chr(39), chr(39) * 2)}'"

    @classmethod
    def quote_sheet_ref(cls, sheet_name, range_ref):
        return f"{cls.quote_sheet_name(sheet_name)}!{range_ref}"

    def add_defined_name(self, name, sheet_name, range_ref):
        """range_ref is the cell range without the sheet, e.g. '$A$4:$A$203'."""
        from openpyxl.workbook.defined_name import DefinedName

        attr_text = self.quote_sheet_ref(sheet_name, range_ref)
        self.wb.defined_names[name] = DefinedName(name=name, attr_text=attr_text)
        return attr_text

    # ----- SAVE -----

    def save(self, path):
        for ws in self.wb.worksheets:
            self._sanitize_sheet_view(ws)
            self._resolve_chart_overlaps(ws)
            self._warn_chart_data_overlap(ws)
        self.wb.save(path)
        return path

    @staticmethod
    def _sanitize_sheet_view(ws):
        """Excel logs a "Repaired Records: View" warning when
        <selection pane="bottomLeft"> exists without a <pane> element.
        Drop the orphan pane attribute."""
        if not ws.sheet_view.pane:
            for sel in ws.sheet_view.selection:
                sel.pane = None

    @staticmethod
    def _chart_span(chart):
        """Default col ≈ 2.3cm, row ≈ 0.5cm."""
        f = chart.anchor._from
        return (
            f.col,
            f.row,
            f.col + max(1, math.ceil((chart.width or 15) / 2.3)),
            f.row + max(1, math.ceil((chart.height or 7.5) * 2)),
        )

    @classmethod
    def _resolve_chart_overlaps(cls, ws):
        placed = []
        for c in ws._charts:
            x0, y0, x1, y1 = cls._chart_span(c)
            for p in placed:
                px0, py0, px1, py1 = cls._chart_span(p)
                if x0 < px1 and px0 < x1 and y0 < py1 and py0 < y1:
                    dx = px0 - c.anchor._from.col
                    dy = (py1 + 1) - c.anchor._from.row
                    c.anchor._from.col += dx
                    c.anchor._from.row += dy
                    if getattr(c.anchor, "to", None) is not None:
                        c.anchor.to.col += dx
                        c.anchor.to.row += dy
                    x0, y0, x1, y1 = cls._chart_span(c)
            placed.append(c)

    @classmethod
    def _warn_chart_data_overlap(cls, ws):
        for idx, c in enumerate(ws._charts, start=1):
            x0, y0, x1, y1 = cls._chart_span(c)
            hits = [
                ws.cell(row=r, column=col).coordinate
                for r in range(y0 + 1, y1 + 1)
                for col in range(x0 + 1, x1 + 1)
                if ws.cell(row=r, column=col).value is not None
            ]
            if hits:
                sample = ", ".join(hits[:5]) + ("..." if len(hits) > 5 else "")
                print(
                    f"[xlsx_skill] WARNING: chart #{idx} on sheet "
                    f"'{ws.title}' covers {len(hits)} populated cell(s): "
                    f"{sample}. Move the chart or relocate the data.",
                    file=sys.stderr,
                )
